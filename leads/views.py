from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import authenticate, login,logout
from .models import Profile,Lead,LeadHistory
from .forms import LeadForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
import openpyxl
from openpyxl.utils import get_column_letter
from django.utils import timezone

def export_leads_excel(leads):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["Name", "Email", "Phone", "Status", "Date Added"]
    ws.append(headers)

    for lead in leads:
        ws.append([
            lead.name,
            lead.email,
            lead.phone,
            lead.status,
            lead.date_added.strftime("%Y-%m-%d %H:%M"),
        ])
    for col in ws.columns:
        max_length = 0
        column = col[0].column

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(column)].width = max_length + 2
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"leads_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response




def login_view(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            role = Profile.objects.get(user=user).role
            if role == "admin":
                return redirect("admin_dashboard")
            else:
                return redirect("employee_dashboard")

        return render(request, "login.html", {"error": "Invalid credentials"})

    return render(request, "login.html")
@login_required
def admin_dashboard(request):
    total_employees = Profile.objects.filter(role="employee").count()
    total_leads = Lead.objects.count()
    new_leads = Lead.objects.filter(status="new").count()
    in_progress = Lead.objects.filter(status="in_progress").count()
    closed = Lead.objects.filter(status="closed").count()
    print("closed",closed)
    context = {
        "total_employees": total_employees,
        "total_leads": total_leads,
        "new_leads": new_leads,
        "in_progress": in_progress,
        "closed_leads": closed
    }

    return render(request,"admin_dashboard.html",context)
@login_required
def lead_list(request):
    profile = request.user.profile
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_leads")
        Lead.objects.filter(id__in=selected_ids).delete()
        return redirect("lead_list")

    search_query = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")
    if profile.role == "admin":
        queryset = Lead.objects.all()
    else:
        queryset = Lead.objects.filter(assigned_to=request.user)
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    

    queryset = queryset.order_by("-id")
    if request.GET.get("export") == "excel":
        return export_leads_excel(queryset)
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get("page")
    leads = paginator.get_page(page_number)
    
    return render(request, "lead_list.html", {
        "leads": leads,
        "search": search_query,
        "status_filter": status_filter,
    })

@login_required
def add_lead(request):
    profile = request.user.profile

    if request.method == "POST":
        form = LeadForm(request.POST, user=request.user)

        if form.is_valid():
            lead = form.save(commit=False)
            if profile.role == "employee":
                lead.assigned_to = request.user
            lead.save()
            messages.success(request,"Lead added successfully")
            return redirect("lead_list")

    else:
        form = LeadForm(user=request.user)

    return render(request, "add_lead.html", {"form": form})

@login_required
def edit_lead(request, pk):
    lead = get_object_or_404(Lead, pk=pk)

    if request.user.profile.role == "employee" and lead.assigned_to != request.user:
        return HttpResponse("Not allowed", status=403)

    if request.method == "POST":
        form = LeadForm(request.POST, instance=lead)

        if request.user.profile.role == "employee":
            form.fields["assigned_to"].disabled = True

        if form.is_valid():
            lead = form.save(commit=False)
            lead.save(user=request.user) 
            messages.success(request,"Lead Updated successfully")
            return redirect("lead_list")
        

    else:
        form = LeadForm(instance=lead)

        if request.user.profile.role == "employee":
            form.fields["assigned_to"].disabled = True

    return render(request, "edit_lead.html", {"form": form})


@login_required
def delete_lead(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if request.user.profile.role == "employee" and lead.assigned_to != request.user:
        return HttpResponse("Not allowed", status=403)
    lead.delete()
    messages.success(request,"Lead deleted successfully")
    return redirect("lead_list")


@login_required
def employee_dashboard(request):
  
    total_leads = Lead.objects.filter(assigned_to=request.user).count()
    print("count",total_leads)
    new_leads = Lead.objects.filter(
        assigned_to=request.user,
        status="new"
    ).count()
    in_progress_leads = Lead.objects.filter(
        assigned_to=request.user,
        status="in_progress"
    ).count()
    closed_leads = Lead.objects.filter(
        assigned_to=request.user,
        status="closed"
    ).count()

    return render(request, "employee_dashboard.html", {
        "total_leads": total_leads,
        "new_leads": new_leads,
        "in_progress": in_progress_leads,
        "closed": closed_leads,
    })

@login_required
def lead_history(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    if request.user.profile.role == "employee" and lead.assigned_to != request.user:
        return HttpResponse("Not allowed", status=403)
    history = LeadHistory.objects.filter(lead=lead).order_by("-updated_at")

    return render(request, "lead_history.html", {"lead": lead, "history": history})
def user_logout(request):
    logout(request)
    return redirect('login')